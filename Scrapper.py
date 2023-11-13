from distutils.log import error
from itertools import count
from numpy import NAN
import pandas as pd
# a  =  [1212,12121111,44455]
# b = [*a]
# print(b)
# data = { 'character' : a }
# frame = pd.DataFrame(data)
# print(frame)
# print(len(a))

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import pickle
import datetime
import pandas as pd
from numpy import NaN
from openpyxl import load_workbook
import re
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support import expected_conditions as EC

driver_path = "geckodriver.exe"
driver = webdriver.Firefox(executable_path=driver_path)

email = "notification.reciever37@gmail.com"
password = "linkedinbot"

# driver.get('https://www.linkedin.com/checkpoint/rm/sign-in-another-account?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin')
# time.sleep(5)
# driver.find_element_by_id('username').send_keys(email)
# time.sleep(2)
# driver.find_element_by_id('password').send_keys(password)
# time.sleep(5)
# driver.find_element_by_id('password').send_keys(Keys.RETURN)
# time.sleep(7)

driver.get("https://www.naukri.com/data-analyst-jobs-in-chennai?k=data%20analyst&l=chennai&experience=5&nignbevent_src=jobsearchDeskGNB")

time.sleep(5)
# with open('cookies.txt', 'wb') as filehandler:
#             pickle.dump(driver.get_cookies(), filehandler)


driver.get("https://www.linkedin.com/jobs/")
time.sleep(5)

with open('cookies.txt', 'rb') as cookiesfile:
    cookies = pickle.load(cookiesfile)
    for cookie in cookies:
        driver.add_cookie(cookie)


# keywords = ['analyst', 'data scientist']
country_name = ['USA']
for country in country_name:
    iteration = 1
    book = load_workbook('job analysis data.xlsx')
    writer = pd.ExcelWriter('job analysis data.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    startrow = writer.sheets['all_data'].max_row
    #startrow=writer.sheets['Sheet1'].max_row
    #0-329=>330 done
    writer.save()
    page_counter = 50
    country_url = f'https://www.linkedin.com/jobs/search/?f_E=1%2C2&f_T=340&keywords=data%20analyst&location={country}&sortBy=R&start={page_counter}'#'https://www.linkedin.com/jobs/search/?f_E=1%2C2&f_T=340&geoId=102713980&keywords=data%20analyst&location=India&sortBy=R'
    time.sleep(1)
    driver.get(country_url)

    time.sleep(6)
    num_jobs = str(driver.find_element_by_class_name('jobs-search-results-list__title-heading').text)


    time.sleep(2)
    jobs = driver.find_elements_by_class_name("occludable-update")
    num_pages = driver.find_elements_by_class_name('artdeco-pagination__indicator')
    position = []
    num_pages = int(num_pages[len(num_pages)-1].text)
    print(f'num_pages = {num_pages}')
    num_page_end = (num_pages-1)*25
    print(f'Page end multiple = {num_page_end}')
    company = []
    location = []
    jd_requirements = []
    date_posted = []
    
    print(f"Starting iterations for {country}.\n Total Jobs for {country} are - {num_jobs}\n Total pages to iterate: {num_pages}\n")
    
    while(page_counter<=num_page_end):
        jobs = driver.find_elements_by_class_name("occludable-update")
        err = 0
        index = 0
        for data in jobs:
            if err == 1:
                time.sleep(6)
                driver.get(f'https://www.linkedin.com/jobs/search/?f_E=1%2C2&f_T=340&keywords=data%20analyst&location={country}&sortBy=R&start={page_counter}')
                time.sleep(6)
                index += 1
                iteration += 1
                jobs = driver.find_elements_by_class_name("occludable-update")
                # driver.execute_script("arguments[0].scrollIntoView();", jobs[index])
                time.sleep(8)
                err = 0
            print(f"On iteration {index} of page {page_counter/25 + 1}")
            try:
                job = jobs[index]
            except:
                print("ERROR: List index out of range")
                err = 1
                continue
            print(f"Job {iteration} for {country}\n")
            #counter+=1
            driver.execute_script("arguments[0].scrollIntoView();", job)
            job.click()
            time.sleep(3)
            try:
                [p, c, l] = job.text.split('\n')[:3]
                
            except:
                print(f"Job iteration {iteration} skipped.")
                err = 1
                # driver.execute_script("arguments[0].scrollIntoView();", jobs)
                
                continue
                
            
            jd_requirements.append(driver.find_element_by_id("job-details").text)
            
            date_of_job = driver.find_elements_by_class_name('jobs-unified-top-card__posted-date')[0].text
            # print(int(date_of_job.split(' ')[0]))
            # time.sleep(10)
            if 'hour' in date_of_job:
                date_posted.append(datetime.date.today())
            elif 'day' in date_of_job:
                date_posted.append(datetime.date.today() - datetime.timedelta(days = int(date_of_job.split(' ')[0])))
            elif 'month' in date_of_job:
                date_posted.append(datetime.date.today() - datetime.timedelta(days = int(date_of_job.split(' ')[0])*30))
            elif 'week' in date_of_job:
                date_posted.append(datetime.date.today() - datetime.timedelta(days = int(date_of_job.split(' ')[0])*7))
            else:
                date_posted.append(NaN)


            #day_posted.append(datetime.date.today() - datetime.timedelta(days = ))
            position.append(str(p))
            company.append(str(c))
            location.append(str(l))
            iteration += 1
            index += 1
        book = load_workbook('job analysis data.xlsx')
        writer = pd.ExcelWriter('job analysis data.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        
        #print(linkedin_result, '====', number)
        data = { 'position' : position,  'company' : company, 'location' : location, 'date_posted' : date_posted, 'jd_requirements':jd_requirements }
        frame = pd.DataFrame(data)
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]') ##To remove Illegal Characters error while saving the xlsx file.
        frame = frame.applymap(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)
        # for sheetname in writer.sheets: #startrow=writer.sheets['Sheet1'].max_row
        frame.to_excel(writer,sheet_name='all_data', startrow=startrow, index = True,header= False)#frame.to_excel('final_count.xlsx', startrow=192)
        writer.save()
        
        page_counter+=25
        driver.get(f'https://www.linkedin.com/jobs/search/?f_E=1%2C2&f_T=340&keywords=data%20analyst&location={country}&sortBy=R&start={page_counter}')#(f'https://www.linkedin.com/jobs/search/?f_E=1%2C2&f_T=340&geoId=102713980&keywords=data%20analyst&location=India&sortBy=R&start={page_counter}')
        time.sleep(6)
        

        print(len(position),'\n\n\n\n', len(company),'\n\n\n\n', len(location), '\n\n\n\n', len(date_posted), '\n\n\n\n', len(jd_requirements))

# data = { 'position' : position,  'company' : company, 'location' : location, 'date_posted' : date_posted, 'jd_requirements':jd_requirements }
# frame = pd.DataFrame(data)
# frame.to_excel('data.xlsx')



#list_items = [items.text.split('\n')[:3] for items in jobs]

# print(list_items)

# profile = [list_items[element][0] for element in range(0,len(list_items)) if list_items[element][0]]
# company = [list_items[element][1] for element in range(0,len(list_items)) if list_items[element][0]]
# location = [list_items[element][2] for element in range(0,len(list_items)) if list_items[element][0]]
# ##int(wd.find_element_by_css_selector(‘h1>span’).get_attribute(‘innerText’))
# # search_bars = driver.find_element_by_class_name
# print(profile,'\n\n\n\n', company,'\n\n\n\n', location)
time.sleep(5)
driver.close()